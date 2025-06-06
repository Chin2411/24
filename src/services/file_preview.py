from __future__ import annotations

from pathlib import Path
from io import BytesIO
import logging

from PyPDF2 import PdfReader
from pdfminer.high_level import extract_text as pdfminer_extract_text
from docx import Document
from PIL import Image, ImageEnhance, ImageOps
import numpy as np

import pytesseract
import fitz
import cv2

from config import PDF_IMAGE_DPI


logger = logging.getLogger(__name__)


def _deskew_image(img: Image.Image) -> Image.Image:
    """Deskew image using OpenCV."""
    try:
        gray = np.array(img.convert("L"))
        coords = np.column_stack(np.where(gray > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
        (h, w) = gray.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        rotated = cv2.warpAffine(np.array(img), M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        return Image.fromarray(rotated)
    except Exception as exc:  # pragma: no cover - optional dependency
        logger.exception("Deskew failed")
        return img


def _preprocess_image(img: Image.Image) -> Image.Image:
    gray = img.convert("L")
    gray = ImageOps.autocontrast(gray)
    gray = _deskew_image(gray)
    enhancer = ImageEnhance.Contrast(gray)
    gray = enhancer.enhance(2.0)
    bw = gray.point(lambda x: 255 if x > 128 else 0, mode="1")
    bbox = bw.getbbox()
    if bbox:
        gray = gray.crop(bbox)
    return gray


def _ocr_pytesseract(img: Image.Image) -> str:
    texts = []
    for psm in (6, 4, 11, 12):
        try:
            conf = f"--psm {psm}"
            t = pytesseract.image_to_string(img, lang="rus+eng", config=conf)
            texts.append(t)
        except Exception as exc:  # pragma: no cover - unexpected errors
            logger.exception("pytesseract failed for psm %s", psm)
    return max(texts, key=len, default="")


def _ocr_easyocr(img: Image.Image) -> str:
    try:
        import easyocr
        import numpy as np

        reader = easyocr.Reader(["ru", "en"], gpu=False)
        result = reader.readtext(np.array(img))
        return "\n".join(r[1] for r in result)
    except Exception as exc:  # pragma: no cover - optional dependency
        logger.exception("EasyOCR error")
    return ""


def _ocr_paddle(img: Image.Image) -> str:
    try:
        from paddleocr import PaddleOCR
        import numpy as np

        ocr = PaddleOCR(use_angle_cls=True, lang="ru")
        result = ocr.ocr(np.array(img), cls=True)
        lines = [line[1][0] for line in result[0]]
        return "\n".join(lines)
    except Exception as exc:  # pragma: no cover - optional dependency
        logger.exception("PaddleOCR error")
    return ""


def _ocr_table_cells(img: Image.Image) -> str:
    """Split table image into cells and run pytesseract on each."""
    cv_img = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
    _, thresh = cv2.threshold(cv_img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(1, img.height // 100)))
    hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(1, img.width // 100), 1))
    vert_lines = cv2.dilate(cv2.erode(thresh, vert_kernel, iterations=1), vert_kernel, iterations=1)
    hor_lines = cv2.dilate(cv2.erode(thresh, hor_kernel, iterations=1), hor_kernel, iterations=1)
    grid = cv2.bitwise_and(vert_lines, hor_lines)
    contours, _ = cv2.findContours(grid, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    boxes = [cv2.boundingRect(c) for c in contours if cv2.contourArea(c) > 100]
    boxes.sort(key=lambda b: (b[1], b[0]))
    rows: list[list[tuple[int, int, int, int]]] = []
    current_y = -1
    row: list[tuple[int, int, int, int]] = []
    for box in boxes:
        x, y, w, h = box
        if current_y == -1 or abs(y - current_y) <= h // 2:
            row.append(box)
            current_y = y
        else:
            rows.append(sorted(row, key=lambda b: b[0]))
            row = [box]
            current_y = y
    if row:
        rows.append(sorted(row, key=lambda b: b[0]))
    lines = []
    for row_boxes in rows:
        cell_texts = []
        for x, y, w, h in row_boxes:
            cell_img = img.crop((x, y, x + w, y + h))
            txt = _ocr_pytesseract(cell_img).strip().replace("\n", " ")
            cell_texts.append(txt)
        lines.append(",".join(cell_texts))
    return "\n".join(lines)


def _extract_tables(path: Path, page: int = 0) -> str:
    """Extract tables from the given PDF page."""
    tables_text = ""
    try:
        import camelot

        tables = camelot.read_pdf(str(path), pages=str(page + 1))
        for table in tables:
            tables_text += table.df.to_csv(index=False) + "\n"
        if tables_text:
            return tables_text
        logger.info("Camelot returned no tables")
    except Exception as exc:  # pragma: no cover - optional dependency
        logger.exception("Camelot error page %s for %s", page, path)
    try:
        import pdfplumber
        import pandas as pd

        with pdfplumber.open(str(path)) as pdf:
            if page < len(pdf.pages):
                p = pdf.pages[page]
                for tbl in p.extract_tables() or []:
                    df = pd.DataFrame(tbl[1:], columns=tbl[0])
                    tables_text += df.to_csv(index=False) + "\n"
        if tables_text:
            return tables_text
        logger.info("pdfplumber returned no tables")
    except Exception as exc:  # pragma: no cover - optional dependency
        logger.exception("pdfplumber error page %s for %s", page, path)

    # --- Fallback using OpenCV cell detection ----------------------------
    try:
        tables_text = _extract_tables_cv(path, page)
        if tables_text:
            return tables_text
    except Exception as exc:  # pragma: no cover - unexpected errors
        logger.exception("OpenCV table extraction failed page %s for %s", page, path)
    return ""


def _extract_tables_cv(path: Path, page: int = 0) -> str:
    """Extract table text from a PDF page using OpenCV."""
    tables_text = ""
    doc = fitz.open(str(path))
    if page < len(doc):
        p = doc[page]
        pix = p.get_pixmap(dpi=PDF_IMAGE_DPI)
        img = Image.open(BytesIO(pix.tobytes()))
        img = _preprocess_image(img)
        text = _ocr_table_cells(img)
        if text.strip():
            tables_text += text + "\n"
    doc.close()
    return tables_text


def _pdf_preview(path: Path) -> tuple[str, str | None]:
    """Return text preview for the first page of a PDF.

    The function tries multiple extraction methods in the following order:
    1. ``PyPDF2`` text layer extraction for the first page.
    2. ``pdfminer.six`` extraction if PyPDF2 returned no text.
    3. ``PyMuPDF`` OCR for scanned pages using ``pytesseract``.

    Returns tuple ``(text, image_path)``. ``image_path`` is provided when text
    couldn't be extracted.
    """

    logger.info("Извлечение превью PDF: %s", path)
    page_num = 0
    MAX_CHARS = 2000

    last_error = ""

    # --- Try PyPDF2 ------------------------------------------------------
    try:
        reader = PdfReader(str(path))
        page = reader.pages[page_num]
        text = (page.extract_text() or "")[:MAX_CHARS]
        if len(text.strip()) > 50:
            return text.strip(), None
        last_error = "PyPDF2 не нашёл текст"
    except Exception as exc:  # pragma: no cover - unexpected errors
        last_error = f"Ошибка PyPDF2: {exc}"
        logger.exception(last_error)

    # --- Try pdfminer ----------------------------------------------------
    try:
        text = pdfminer_extract_text(str(path), page_numbers=[page_num]) or ""
        text = text[:MAX_CHARS]
        if len(text.strip()) > 50:
            return text.strip(), None
        last_error = "pdfminer не нашёл текст"
    except Exception as exc:  # pragma: no cover - unexpected errors
        last_error = f"Ошибка pdfminer: {exc}"
        logger.exception(last_error)

    pages = str(page_num + 1)

    # --- OCR fallback using PyMuPDF + pytesseract -----------------------
    text = ""
    try:
        doc = fitz.open(str(path))
        if page_num < len(doc):
            page = doc[page_num]
            pix = page.get_pixmap(dpi=PDF_IMAGE_DPI)
            img = Image.open(BytesIO(pix.tobytes()))
            try:
                Path("logs").mkdir(exist_ok=True)
                img.save(Path("logs") / f"{path.stem}_page1.png")
            except Exception as exc:  # pragma: no cover - optional
                logger.exception("Failed to save debug image")
            img = _preprocess_image(img)
            ocr_text = _ocr_pytesseract(img)
            if len(ocr_text.strip()) < 20:
                ocr_text = _ocr_easyocr(img)
            if len(ocr_text.strip()) < 20:
                ocr_text = _ocr_paddle(img)
            text += ocr_text
            if len(text) >= MAX_CHARS:
                text = text[:MAX_CHARS]
        doc.close()
    except Exception as exc:  # pragma: no cover - unexpected errors
        last_error = f"Ошибка OCR: {exc}"
        logger.exception(last_error)

    if len(text.strip()) < 50:
        table_text = _extract_tables(path, page_num)
        if not table_text:
            table_text = _extract_tables_cv(path, page_num)
        text += table_text

    if text.strip():
        logger.info("Превью PDF создано для %s", path)
        return text.strip(), None

    image_path = None
    try:
        doc = fitz.open(str(path))
        pix = doc[page_num].get_pixmap(dpi=PDF_IMAGE_DPI)
        image_path = str(Path("logs") / f"{path.stem}_preview.png")
        Path(image_path).parent.mkdir(exist_ok=True)
        Image.open(BytesIO(pix.tobytes())).save(image_path)
        doc.close()
    except Exception as exc:  # pragma: no cover - unexpected errors
        logger.exception("Failed to save preview image")

    logger.error(
        "Не удалось корректно распознать таблицу: %s", last_error or "unknown"
    )
    logger.info("Возвращено изображение превью для %s", path)
    return "", image_path


def _docx_preview(path: Path) -> str:
    logger.info("Извлечение превью DOCX: %s", path)
    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs[:25])
    result = text.strip()[:2000]
    logger.info("Превью DOCX создано для %s", path)
    return result


def _text_preview(path: Path) -> str:
    logger.info("Извлечение превью текста: %s", path)
    lines: list[str] = []
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            for _ in range(20):
                line = f.readline()
                if not line:
                    break
                lines.append(line)
    except Exception as exc:
        raise RuntimeError(f"Ошибка чтения файла: {exc}")
    result = "".join(lines).strip()
    logger.info("Превью текста создано для %s", path)
    return result


def _xlsx_preview(path: Path, rows: int = 30) -> str:
    """Return text preview of the first ``rows`` rows of the first sheet.

    ``openpyxl`` is used directly to avoid heavy pandas import and to work
    reliably even when ``xlrd`` is not installed. Only the first worksheet is
    read for performance reasons.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(f"Missing openpyxl: {exc}") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    preview_lines: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        values = ["" if v is None else str(v) for v in row]
        preview_lines.append("\t".join(values))
        if i + 1 >= rows:
            break
    wb.close()
    return "\n".join(preview_lines)


def _excel_preview(path: Path) -> str:
    """Return CSV preview for Excel files."""
    logger.info("Извлечение превью Excel: %s", path)
    try:
        import pandas as pd

        df = pd.read_excel(path, nrows=20)
        result = df.to_csv(index=False)
        logger.info("Превью Excel создано для %s", path)
        return result
    except Exception as exc:
        logger.exception("Ошибка чтения Excel %s", path)
        raise RuntimeError(f"Ошибка чтения Excel: {exc}")


def _image_preview(path: Path) -> str:
    logger.info("Извлечение превью изображения: %s", path)
    try:
        with Image.open(path) as img:
            img = _preprocess_image(img)
            text = _ocr_pytesseract(img)
            if len(text.strip()) < 20:
                text = _ocr_easyocr(img)
            if len(text.strip()) < 20:
                text = _ocr_paddle(img)
            if text.strip():
                result = text.strip()
                logger.info("OCR изображение завершено %s", path)
                return result
            Path("logs").mkdir(exist_ok=True)
            img.save(Path("logs") / f"{Path(path).stem}_debug.png")
            raise RuntimeError(
                "OCR не справился, возможно, сложная таблица или качество недостаточно"
            )
    except Exception as exc:
        logger.exception("Ошибка OCR в файле %s", path)
        raise RuntimeError(f"Ошибка OCR: {exc}")


SUPPORTED_IMAGES = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}
SUPPORTED_TEXT = {".txt", ".csv"}
SUPPORTED_EXCEL = {".xlsx", ".xls"}
SUPPORTED_DOCS = {".docx", ".doc"}


def extract_preview(path: Path) -> tuple[str, str | None, str]:
    """Return preview text or image for supported file types."""
    logger.info("Извлечение превью файла: %s", path)
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            text, image = _pdf_preview(path)
            logger.info("Превью PDF готово %s", path)
            return text, image, ""
        if ext in SUPPORTED_DOCS:
            text = _docx_preview(path)
            logger.info("Превью DOCX готово %s", path)
            return text, None, ""
        if ext in SUPPORTED_TEXT:
            text = _text_preview(path)
            logger.info("Превью текста готово %s", path)
            return text, None, ""
        if ext in SUPPORTED_EXCEL:
            text = _xlsx_preview(path)
            logger.info("Превью Excel готово %s", path)
            return text, None, ""
        if ext in SUPPORTED_IMAGES:
            text = _image_preview(path)
            logger.info("Превью изображения готово %s", path)
            return text, None, ""
    except Exception as exc:
        return "", None, str(exc)
    return "", None, "Просмотр не поддерживается"


def extract_preview_text(path: Path) -> str:
    logger.info("Извлечение текстового превью из файла: %s", path)
    text, _, _ = extract_preview(path)
    logger.info("Возвращено текстовое превью для %s", path)
    return text
